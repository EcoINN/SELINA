import os
from google.cloud import vision
from openpyxl import load_workbook


def analyze_images_from_excel(excel_path, credentials_path, column_name="url_1"):
    """
    Analyse images from a given Excel file using the Google Cloud Vision API.

    This function reads URLs from a specified column in the Excel file, sends these URLs
    to the Vision API for label detection, and writes the detected labels in the column
    immediately to the right of the specified one.

    Parameters:
    - excel_path (str): The path to the Excel file containing the image URLs.
    - credentials_path (str): The path to the Google Cloud JSON credentials file.
    - column_name (str, optional): The name of the column containing the image URLs.
                                   Defaults to "url_1".

    Returns:
    None. The function updates the provided Excel file in place.

    Usage:
    ```python
    analyze_images_from_excel("path\to\your\excel\file.xlsx", "path\to\SELINA\your-credentials.json", "url_1")
    ```
    """
    # Set the credentials for the session
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = credentials_path

    # Load the Excel file
    wb = load_workbook(excel_path)
    ws = wb.active

    # Get the image URLs from the specified column
    image_urls = [cell.value for cell in ws[column_name] if cell.value]

    client = vision.ImageAnnotatorClient()

    for idx, url in enumerate(image_urls, start=1):
        image = vision.Image()
        image.source.image_uri = url
        response = client.label_detection(image=image)
        labels = [label.description for label in response.label_annotations]

        # Write the labels to the Excel file in the column next to the specified one
        # Assuming there are no gaps in the columns and they are consecutive (e.g., A, B, C, ...)
        next_column = chr(ord(column_name[0]) + 1)
        ws[f'{next_column}{idx}'] = ', '.join(labels)

    # Save the updated Excel file
    wb.save(excel_path)
    print(f"Analysis completed. Results saved in {excel_path}.")


test = analyze_images_from_excel("C:\Ecostack\02_Projects\01_Selina\selina\output\Mt_tweets.xlsx", "C:\Ecostack\02_Projects\01_Selina\vision\selinadp-88bc54370d2c.json", "url-1")